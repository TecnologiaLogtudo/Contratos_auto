import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox
import threading
import queue
import os
import configparser
import openpyxl
from datetime import datetime
from typing import Optional
import sys
import time

# Importa a lógica das fases
import fase1_processamento
import fase2_login
import fase3_preenchimento
import fase4_frete
from playwright.sync_api import Error as PlaywrightError
import fase5_contrato_frete

# --- Função para encontrar o caminho correto dos recursos (para PyInstaller) ---
def get_resource_path(relative_path):
    """ Retorna o caminho absoluto para um recurso, funcionando em dev e no PyInstaller. """
    if getattr(sys, 'frozen', False):
        # Se rodando como um executável, o caminho base para recursos é _MEIPASS
        base_path = sys._MEIPASS
    else:
        # Se rodando como script, o caminho base é o diretório do arquivo
        base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)

def get_persistent_path(filename):
    """ Retorna um caminho na pasta do executável para arquivos persistentes (logs, configs). """
    if getattr(sys, 'frozen', False):
        # Pasta do executável
        return os.path.join(os.path.dirname(sys.executable), filename)
    # Pasta do script em desenvolvimento
    return os.path.join(os.path.dirname(__file__), filename)

# --- Constantes ---
APP_TITLE = "Automação de Contratos LogTudo v2.0"
LOG_FILENAME = get_persistent_path("log_automacao.txt")
CONFIG_FILENAME = get_persistent_path("config.ini")
URL_DESTINO = "https://logtudo.e-login.net/versoes/versao5.0/rotinas/formulario.php?rotina=trans_conhecimento&OP=O1&_qsf=1"

class AutomacaoUI:
    """
    Classe principal da Interface Gráfica (UI) com Tkinter.
    """
    
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("850x650")
        self.root.minsize(600, 400)

        self.config = configparser.ConfigParser()
        self.config_path = CONFIG_FILENAME # Já é o caminho absoluto correto

        self.planilha_processada_path: Optional[str] = None
        self.playwright_instance = None
        self.browser = None
        self.page = None
        self.log_queue = queue.Queue()

        self.pause_event = threading.Event()
        self.pause_event.set()
        self.is_paused = False
        self.previous_tab_index = 0 # Adicionado para rastrear a aba anterior

        # Define o ícone da janela
        icon_path = get_resource_path(os.path.join("assets", "app_icon.ico"))
        self.root.iconbitmap(default=icon_path)
        
        self.var_login = tk.StringVar()
        self.var_senha = tk.StringVar()
        self.var_atraso_fases = tk.StringVar(value='1')
        self.var_atraso_etapas = tk.StringVar(value='0.3')
        self.var_dados_km = tk.StringVar(value='10')

        self.criar_widgets()
        self.iniciar_processador_log()
        self.carregar_configuracoes()

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def criar_widgets(self):
        """Cria todos os elementos visuais da interface."""
        
        # --- Barra de Status (criada primeiro para ficar no fundo) ---
        self.status_var = tk.StringVar(value="Pronto")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor="w", padding=5)
        status_bar.pack(side="bottom", fill="x")

        # --- Notebook para Abas (agora o elemento principal da janela) ---
        # Removido o padding de 10px para eliminar o espaço extra
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Adiciona um log para a interação do usuário com as abas
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

        # --- Aba de Execução ---
        frame_execucao = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(frame_execucao, text="Execução")
        frame_execucao.grid_rowconfigure(1, weight=1) # A área de logs expande
        frame_execucao.grid_columnconfigure(0, weight=1)

        # Frame para os botões de controle
        frame_controles = ttk.Frame(frame_execucao)
        frame_controles.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        frame_controles.grid_columnconfigure(0, weight=1) # Botões se expandem

        self.btn_iniciar = ttk.Button(
            frame_controles,
            text="▶ Iniciar Automação",
            command=self.iniciar_automacao_completa,
            style="Accent.TButton"
        )
        self.btn_iniciar.grid(row=0, column=0, padx=5, pady=5, sticky="ew")

        self.btn_pause = ttk.Button(
            frame_controles,
            text="Pausar",
            command=self.toggle_pause,
            state="disabled"
        )
        self.btn_pause.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        # Área de Logs
        self.log_area = scrolledtext.ScrolledText(
            frame_execucao,
            wrap=tk.WORD, 
            state="disabled", 
            height=15,
            font=("Consolas", 9)
        ) 
        self.log_area.grid(row=1, column=0, sticky="nsew")

        # Configuração das cores do log
        self.log_area.tag_config("INFO", foreground="#000000")
        self.log_area.tag_config("DEBUG", foreground="#808080")
        self.log_area.tag_config("AVISO", foreground="#FFA500")
        self.log_area.tag_config("ERRO", foreground="#FF0000", font=("Consolas", 9, "bold"))
        self.log_area.tag_config("SUCESSO", foreground="#008000", font=("Consolas", 9, "bold"))
        self.log_area.tag_config("FASE", foreground="#0000FF", font=("Consolas", 9, "bold"))

        # --- Aba de Configurações ---
        frame_config = ttk.Frame(self.notebook, padding=20)
        self.notebook.add(frame_config, text="Configurações")
        frame_config.grid_columnconfigure(0, weight=1) # Campos se expandem

        # Grupo de Credenciais
        lf_credenciais = ttk.LabelFrame(frame_config, text="Credenciais de Acesso", padding=15)
        lf_credenciais.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        lf_credenciais.grid_columnconfigure(1, weight=1)

        lbl_login = ttk.Label(lf_credenciais, text="Login:")
        lbl_login.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ent_login = ttk.Entry(lf_credenciais, textvariable=self.var_login, width=40)
        ent_login.grid(row=0, column=1, sticky="ew", padx=5, pady=5)

        lbl_senha = ttk.Label(lf_credenciais, text="Senha:")
        lbl_senha.grid(row=1, column=0, sticky="w", padx=5, pady=5)
        ent_senha = ttk.Entry(lf_credenciais, textvariable=self.var_senha, show="*", width=40)
        ent_senha.grid(row=1, column=1, sticky="ew", padx=5, pady=5)

        # Grupo de Parâmetros
        lf_params = ttk.LabelFrame(frame_config, text="Parâmetros de Execução", padding=15)
        lf_params.grid(row=1, column=0, sticky="ew")
        lf_params.grid_columnconfigure(1, weight=1)

        lbl_atraso_fases = ttk.Label(lf_params, text="Atraso entre fases (s):")
        lbl_atraso_fases.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        ent_atraso_fases = ttk.Entry(lf_params, textvariable=self.var_atraso_fases, width=10)
        ent_atraso_fases.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        lbl_atraso_etapas = ttk.Label(lf_params, text="Atraso entre etapas (s):")
        lbl_atraso_etapas.grid(row=1, column=0, sticky="w", padx=5, pady=5)
        ent_atraso_etapas = ttk.Entry(lf_params, textvariable=self.var_atraso_etapas, width=10)
        ent_atraso_etapas.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        lbl_dados_km = ttk.Label(lf_params, text="Km Padrão:")
        lbl_dados_km.grid(row=2, column=0, sticky="w", padx=5, pady=5)
        ent_dados_km = ttk.Entry(lf_params, textvariable=self.var_dados_km, width=10)
        ent_dados_km.grid(row=2, column=1, sticky="w", padx=5, pady=5)

        # Estilo do botão de destaque
        ttk.Style().configure("Accent.TButton", font=("Segoe UI", 10, "bold"), foreground="black", background="#0078D7")

    def carregar_configuracoes(self):
        """Lê o config.ini e preenche os campos da UI."""
        try:
            if os.path.exists(self.config_path):
                self.config.read(self.config_path, encoding='utf-8')
                if 'CREDENCIAS' in self.config:
                    self.var_login.set(self.config['CREDENCIAS'].get('Login', ''))
                    self.var_senha.set(self.config['CREDENCIAS'].get('Senha', ''))
                if 'AUTOMACAO' in self.config:
                    self.var_atraso_fases.set(self.config['AUTOMACAO'].get('AtrasoFases', '1'))
                    self.var_atraso_etapas.set(self.config['AUTOMACAO'].get('AtrasoEtapas', '0.3'))
                    self.var_dados_km.set(self.config['AUTOMACAO'].get('Dados_km', '10'))
            else:
                self.log("Arquivo config.ini não encontrado. Criando um novo.", "AVISO")
                self.config['CREDENCIAS'] = {'Login': '', 'Senha': ''}
                self.config['AUTOMACAO'] = {'AtrasoFases': '1', 'AtrasoEtapas': '0.3', 'Dados_km': '10'}
                self.salvar_configuracoes()

        except Exception as e:
            self.log(f"Erro ao carregar configurações: {e}", "ERRO")

    def salvar_configuracoes(self):
        """Salva os campos da UI no config.ini."""
        try:
            if 'CREDENCIAS' not in self.config:
                self.config['CREDENCIAS'] = {}
            if 'AUTOMACAO' not in self.config:
                self.config['AUTOMACAO'] = {}
                
            self.config['CREDENCIAS']['Login'] = self.var_login.get()
            self.config['CREDENCIAS']['Senha'] = self.var_senha.get()
            self.config['AUTOMACAO']['AtrasoFases'] = self.var_atraso_fases.get()
            self.config['AUTOMACAO']['AtrasoEtapas'] = self.var_atraso_etapas.get()
            self.config['AUTOMACAO']['Dados_km'] = self.var_dados_km.get()
            
            with open(self.config_path, 'w', encoding='utf-8') as configfile:
                self.config.write(configfile)
            self.log("Configurações salvas.", "DEBUG")
            
        except Exception as e:
            self.log(f"Erro ao salvar configurações: {e}", "ERRO")

    def log(self, mensagem: str, tipo: str = "INFO", espacamento_antes: bool = False):
        """Envia uma mensagem de log para a fila."""
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Adiciona uma linha em branco antes da mensagem, se solicitado
            prefixo_espaco = "\n" if espacamento_antes else ""
            log_formatado = f"{prefixo_espaco}[{timestamp}] [{tipo.upper()}]: {mensagem}\n"
            
            with open(LOG_FILENAME, "a", encoding="utf-8") as f:
                f.write(log_formatado)
            
            self.log_queue.put((f"[{timestamp}] [{tipo.upper()}]: {mensagem}\n", tipo, espacamento_antes))
            
        except Exception as e:
            print(f"ERRO NO SISTEMA DE LOG: {e}")

    def iniciar_processador_log(self):
        """Inicia o loop que consome a fila de log e atualiza a UI."""
        try:
            while not self.log_queue.empty():
                mensagem, tipo, espacamento_antes = self.log_queue.get_nowait()
                
                self.log_area.configure(state="normal")
                if espacamento_antes:
                    # Adiciona a linha em branco na área de log da UI
                    self.log_area.insert(tk.END, "\n")
                self.log_area.insert(tk.END, mensagem, tipo)
                self.log_area.configure(state="disabled")
                self.log_area.see(tk.END)
            
            self.root.after(100, self.iniciar_processador_log)
            
        except queue.Empty:
            self.root.after(100, self.iniciar_processador_log)
        except Exception as e:
            print(f"Erro no processador de log da UI: {e}")
            self.root.after(100, self.iniciar_processador_log)

    def on_tab_changed(self, event):
        """Loga a mudança de aba e salva as configurações ao sair da aba de config."""
        try:
            selected_tab_index = self.notebook.index(self.notebook.select())
            tab_text = self.notebook.tab(selected_tab_index, "text")
            self.log(f"Usuário navegou para a aba '{tab_text}'.", "INFO")
    
            # Salva as configurações automaticamente ao sair da aba "Configurações" (índice 1)
            if self.previous_tab_index == 1 and selected_tab_index != 1:
                self.salvar_configuracoes()
    
            self.previous_tab_index = selected_tab_index
        except Exception as e:
            self.log(f"Erro ao processar mudança de aba: {e}", "ERRO")

    def on_closing(self):
        """Chamado quando o 'X' da janela é clicado."""
        if messagebox.askokcancel("Sair", "Deseja realmente sair?"):
            self.log("Fechando aplicação...", "INFO")
            
            if self.is_paused:
                self.toggle_pause()
            
            self.salvar_configuracoes()
            
            self.btn_iniciar.config(state="disabled")
            self.btn_pause.config(state="disabled", text="Pausar")

            if self.browser or self.playwright_instance:
                threading.Thread(target=self.tarefa_fechar_browser, daemon=True).start()
            else:
                self.root.destroy()

    def tarefa_fechar_browser(self):
        """Função executada em thread para fechar o Playwright."""
        try:
            if self.browser:
                self.browser.close()
            if self.playwright_instance:
                self.playwright_instance.stop()
        except Exception as e:
            self.log(f"Erro ao fechar o navegador: {e}", "ERRO")
        finally:
            self.root.after(0, self.root.destroy)

    def toggle_pause(self):
        """Alterna o estado de pausa da automação."""
        if self.is_paused:
            self.is_paused = False
            self.pause_event.set()
            self.btn_pause.config(text="Pausar")
            self.log("Automação retomada.", "INFO")
            self.status_var.set("Executando...")
        else:
            self.is_paused = True
            self.pause_event.clear()
            self.btn_pause.config(text="Continuar")
            self.log("Automação pausada. Clique 'Continuar' para retomar.", "AVISO")
            self.status_var.set("Pausado")

    def iniciar_automacao_completa(self):
        """Inicia o fluxo completo da automação em uma thread."""
        self.log("Botão 'Iniciar Automação' clicado. Abrindo seletor de arquivo...", "INFO")
        
        filepath = filedialog.askopenfilename(
            title="Selecione a planilha (.xlsx ou .xls)",
            filetypes=(
                ("Planilhas Excel", "*.xlsx *.xls"),
                ("Todos os arquivos", "*.*")
            )
        )
        if not filepath:
            self.log("Seleção de arquivo cancelada.", "AVISO")
            return

        self.log(f"Arquivo selecionado: {filepath}", "INFO")

        login = self.var_login.get()
        senha = self.var_senha.get()
        if not login or not senha:
            self.log("Erro: Login e Senha devem ser preenchidos.", "ERRO")
            self.notebook.select(self.frame_config)
            return

        self.btn_iniciar.config(state="disabled", text="Automação em Andamento...")
        self.btn_pause.config(state="normal")
        self.is_paused = False
        self.pause_event.set()
        self.btn_pause.config(text="Pausar")

        threading.Thread(
            target=self.tarefa_automacao_completa,
            args=(filepath, login, senha),
            daemon=True
        ).start()

    def _reset_session_by_relogging(self, nro_cotacao: str, atraso_fases: float, force_new_page: bool = False):
        """
        Força um reset completo da sessão realizando logout e login novamente.
        Esta é uma abordagem robusta para limpar formulários persistentes.
        Reutiliza a página existente para evitar múltiplas janelas, a menos que force_new_page seja True.
        """
        # Se force_new_page for True, o perform_login criará uma nova página.
        # Isso pode ser útil em cenários de erro mais graves onde a página atual está corrompida.
        self.log(f"[F0] [Item {nro_cotacao}] Iniciando reset de sessão via logout/login...", "DEBUG")
        try:
            # --- ETAPA 1: Logout ---
            self.log(f"[F0] [Item {nro_cotacao}] Passo 1/3: Realizando logout...", "DEBUG")
            
            # Localiza o botão do menu do usuário e passa o mouse sobre ele
            try:
                user_menu_button = self.page.locator("button.dropbtn")
                user_menu_button.hover(timeout=10000)
                sair_button = self.page.get_by_text("Sair", exact=True)
                with self.page.expect_navigation(wait_until="networkidle", timeout=30000):
                    sair_button.click()
                
                # --- VERIFICAÇÃO DE LOGOUT ---
                # Após o clique em "Sair", a página deve navegar para a URL de login.
                # Podemos verificar a URL ou a presença de um elemento da página de login.
                if "logtudo.e-login.net" in self.page.url:
                    self.log(f"[F0] [Item {nro_cotacao}] Logout realizado com sucesso e página de login detectada.", "DEBUG")
                else:
                    # Se a URL não for a esperada, pode ser que o logout não tenha ocorrido ou a navegação falhou.
                    raise PlaywrightError(f"Logout falhou: Página atual '{self.page.url}' não é a página de login esperada.")
            except PlaywrightError as e:
                self.log(f"[F0] [Item {nro_cotacao}] ERRO: Não foi possível realizar logout ou verificar o sucesso: {e}. Tentando login direto.", "ERRO")
                # Se o logout falhar, a página pode já estar na tela de login ou em um estado inesperado.
                # Prosseguimos para o login, que tentará navegar para a URL de login de qualquer forma.

            # --- ETAPA 2: Novo Login ---
            self.log(f"[F0] [Item {nro_cotacao}] Passo 2/3: Realizando novo login na página existente...", "DEBUG")
            login = self.var_login.get()
            senha = self.var_senha.get()
            # Passa a página existente para a função de login
            self.page = fase2_login.perform_login(self.browser, login, senha, URL_DESTINO, self.log, existing_page=self.page if not force_new_page else None)
            if not self.page:
                raise Exception("Falha ao realizar novo login durante o reset da sessão.")
            self.log(f"[F0] [Item {nro_cotacao}] Novo login e navegação para URL de destino concluídos.", "SUCESSO")

        except Exception as e:
            self.log(f"[F0] [Item {nro_cotacao}] ERRO CRÍTICO durante o reset da sessão: {e}", "ERRO")
            self.log(f"[F0] [Item {nro_cotacao}] A automação será interrompida para evitar mais erros.", "ERRO")
            # Lançar a exceção para que o loop principal pare
            raise e

    def _is_planilha_tratada(self, filepath: str) -> bool:
        """
        Verifica se a planilha selecionada já está no formato processado,
        comparando os cabeçalhos da primeira aba.
        """
        try:
            self.log(f"[F0] Verificando se a planilha '{os.path.basename(filepath)}' já foi tratada...", "DEBUG")
            workbook = openpyxl.load_workbook(filepath, read_only=True)
            sheet = workbook.active
            
            headers_encontrados = [cell.value for cell in sheet[1]]
            # CORREÇÃO: A lista de cabeçalhos esperados deve corresponder exatamente à criada na Fase 1.
            headers_esperados = ["Nro cotação", "Categoria veículo", "Cidade", "UF", "Nome", "Placa", "Data pagamento", "Viagem extra", "Remetente", "Status"]
            
            # Compara os cabeçalhos (ignorando células extras vazias)
            headers_limpos = [h for h in headers_encontrados if h is not None]
            
            # A verificação agora é simples: os cabeçalhos limpos devem ser iguais aos esperados.
            return headers_limpos == headers_esperados
        except PermissionError as e:
            # Erro específico de permissão. É importante parar aqui.
            self.log(f"[F0] ERRO DE PERMISSÃO ao verificar a planilha '{os.path.basename(filepath)}'.", "ERRO")
            self.log("[F0] Verifique se o arquivo já está aberto no Excel ou em outro programa. Feche-o e tente novamente.", "ERRO")
            # Lança a exceção para que o fluxo principal possa capturá-la e parar a automação.
            raise e
        except Exception as e:
            self.log(f"[F0] Não foi possível verificar a planilha. Assumindo que não está tratada. Erro: {e}", "AVISO")
            # Se houver qualquer erro na leitura, assume que a planilha não está tratada.
            return False


    def _ler_dados_planilha(self) -> Optional[list]:
        """Helper para ler os dados da planilha processada."""
        if not self.planilha_processada_path:
            self.log("Caminho da planilha processada não encontrado.", "ERRO")
            return None
            
        try:
            self.log(f"Lendo dados de '{os.path.basename(self.planilha_processada_path)}'...", "DEBUG")
            wb = openpyxl.load_workbook(self.planilha_processada_path, data_only=True)
            ws = wb.active
            
            dados = []
            headers = [cell.value for cell in ws[1] if cell.value is not None] 
            self.log(f"Cabeçalhos encontrados: {headers}", "DEBUG")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                
                dados_linha = dict(zip(headers, row))
                
                if dados_linha.get("Nro cotação"):
                    dados.append(dados_linha)
                else:
                    self.log(f"Linha ignorada (sem Nro cotação): {dados_linha}", "DEBUG")

            self.log(f"Total de {len(dados)} itens lidos da planilha.", "INFO")
            return dados
            
        except Exception as e:
            self.log(f"Erro ao ler arquivo Excel: {e}", "ERRO")
            return None

    def tarefa_automacao_completa(self, filepath: str, login: str, senha: str):
        """Executa todo o fluxo de automação: Fase 1, Fase 2 e o loop das Fases 3-5."""
        self.status_var.set("Executando: Fase 1 - Processando planilha...")
        try:
            # --- VERIFICAÇÃO E FASE 1 ---
            if self._is_planilha_tratada(filepath):
                self.log("[F1] Planilha já tratada detectada. Pulando Fase 1.", "INFO", espacamento_antes=True)
                self.planilha_processada_path = filepath
            else:
                self.log("[F1] Planilha não tratada detectada. Iniciando processamento...", "INFO", espacamento_antes=True)
                self.log("[F1] --- INÍCIO: PROCESSAMENTO DA PLANILHA ---", "FASE")
                self.planilha_processada_path = fase1_processamento.processar_planilha(filepath, self.log)
                if not self.planilha_processada_path:
                    self.log("Fase 1 falhou. A automação não pode continuar.", "ERRO")
                    # Limpa o status e reabilita os botões no 'finally'
                    return
                self.log("Fase 1 concluída com sucesso.", "SUCESSO")


            # --- Validação de Parâmetros ---
            try:
                atraso_fases = float(self.var_atraso_fases.get())
                if atraso_fases < 0: atraso_fases = 0
            except ValueError:
                self.log(f"Valor de atraso de fases '{self.var_atraso_fases.get()}' é inválido. Usando 1s.", "AVISO")
                atraso_fases = 1.0
            try:
                atraso_etapas = float(self.var_atraso_etapas.get())
                if atraso_etapas < 0: atraso_etapas = 0
            except ValueError:
                self.log(f"Valor de atraso de etapas '{self.var_atraso_etapas.get()}' é inválido. Usando 0.3s.", "AVISO")
                atraso_etapas = 0.3
            
            dados_km = self.var_dados_km.get()
            if not dados_km.isdigit(): # Verifica se é um número
                self.log(f"Valor de Km Padrão '{self.var_dados_km.get()}' é inválido. Usando '10'.", "AVISO")
                dados_km = "10"

            # --- FASE 2 ---
            self.log("[F2] --- INÍCIO: LOGIN E SETUP DO NAVEGADOR ---", "FASE", espacamento_antes=True)
            self.status_var.set("Executando: Fase 2 - Login...")
            
            self.pause_event.set()
            
            if not self.playwright_instance:
                self.playwright_instance, self.browser = fase2_login.launch_browser(self.log)
                if not self.browser:
                    raise Exception("Falha ao iniciar o navegador.")
            
            self.page = fase2_login.perform_login(
                self.browser, login, senha, URL_DESTINO, self.log
            )
            
            if not self.page:
                raise Exception("Falha no login (Fase 2).")

            self.log("Fase 2 (Login) concluída com sucesso.", "SUCESSO")
            
            dados_planilha = self._ler_dados_planilha()
            if not dados_planilha:
                raise Exception("Falha ao ler dados da planilha.")

            # Filtra todos os itens com status "Pendente"
            dados_para_processar = [item for item in dados_planilha if item.get("Status") == "Pendente"]

            if not dados_para_processar:
                self.log("Nenhuma cotação com status 'Pendente' encontrada. A automação será encerrada.", "AVISO")
                return  # Encerra a automação se nenhum item pendente for encontrado

            self.log(f"Encontrados {len(dados_para_processar)} itens com status 'Pendente' para processar.", "INFO")


            # --- LOOP DAS FASES 3-5 ---
            self.log(f"--- INICIANDO PROCESSAMENTO (Fases 3-5) PARA OS {len(dados_para_processar)} ITENS PENDENTES ---", "FASE", espacamento_antes=True)
            
            for i, item_dados in enumerate(dados_para_processar, 1):
                
                self.pause_event.wait() 
                
                nro_cotacao = item_dados.get("Nro cotação", "N/A")
                self.log(f"--- Processando Item {i}/{len(dados_para_processar)} (Cotação: {nro_cotacao}) ---", "INFO", espacamento_antes=True)
                self.status_var.set(f"Executando: Item {i}/{len(dados_para_processar)} (Cotação: {nro_cotacao})")
                
                try:
                    # --- FASE 3: PREENCHIMENTO BÁSICO ---
                    sucesso_fase_3 = fase3_preenchimento.preencher_formulario(
                        self.page, item_dados, self.log, self.pause_event, self.planilha_processada_path, atraso_etapas, atraso_fases
                    )
                    if not sucesso_fase_3:
                        self.log(f"Item {nro_cotacao} falhou na Fase 3. Pulando.", "ERRO")
                        self.log(f"Aguardando {atraso_fases}s antes do próximo item...", "DEBUG")
                        time.sleep(atraso_fases)
                        self._reset_session_by_relogging(nro_cotacao, atraso_fases) # Relogin em caso de erro
                        continue

                    self.log(f"Aguardando {atraso_fases}s antes da próxima fase...", "DEBUG")
                    time.sleep(atraso_fases)

                    # --- FASE 4: PREENCHIMENTO FRETE ---
                    sucesso_fase_4 = fase4_frete.preencher_frete(
                        self.page, item_dados, self.log, self.pause_event, self.planilha_processada_path, atraso_etapas
                    )
                    if not sucesso_fase_4:
                        motivo_erro = f"Falha na Fase 4: Placa '{item_dados.get('Placa', 'N/A')}' não encontrada ou erro no preenchimento do frete."
                        self.log(f"Item {nro_cotacao} falhou na Fase 4. {motivo_erro} Pulando.", "ERRO")
                        fase3_preenchimento.registrar_erro_em_planilha(item_dados, motivo_erro, self.log, self.planilha_processada_path)
                        self.log(f"Aguardando {atraso_fases}s antes do próximo item...", "DEBUG")
                        time.sleep(atraso_fases)
                        self._reset_session_by_relogging(nro_cotacao, atraso_fases) # Relogin em caso de erro
                        continue

                    self.log(f"Aguardando {atraso_fases}s antes da próxima fase...", "DEBUG")
                    time.sleep(atraso_fases)

                    # --- FASE 5: CONTRATO FRETE ---
                    sucesso_fase_5 = fase5_contrato_frete.preencher_contrato_frete(
                        self.page, item_dados, self.log, self.pause_event, atraso_etapas, self.planilha_processada_path, atraso_fases, dados_km
                    )
                    if not sucesso_fase_5:
                        self.log(f"Item {nro_cotacao} falhou na Fase 5. Pulando.", "ERRO")
                        self.log(f"Aguardando {atraso_fases}s antes do próximo item...", "DEBUG")
                        time.sleep(atraso_fases)
                        self._reset_session_by_relogging(nro_cotacao, atraso_fases) # Relogin em caso de erro
                        continue
                    
                    # A Fase 5 retornou sucesso, agora tentamos resetar a página para o próximo item.
                    self.log(f"Aguardando {atraso_fases}s antes de navegar para o próximo item...", "DEBUG")
                    time.sleep(atraso_fases)
                    
                    # Navega para a URL de destino para preparar para o próximo item
                    self.page.goto(URL_DESTINO, wait_until="networkidle", timeout=60000)
                    self.log(f"[F0] [Item {nro_cotacao}] Página de formulário resetada para o próximo item.", "DEBUG")

                except PlaywrightError as e_item:
                    # Erro específico do Playwright
                    error_message = str(e_item)
                    if "Target page, context or browser has been closed" in error_message:
                        self.log("ERRO CRÍTICO: O navegador foi fechado inesperadamente. A automação será interrompida.", "ERRO")
                        self.log("AVISO: Para evitar este erro, não feche a janela do navegador durante a execução.", "AVISO")
                        # Força a saída do loop
                        break
                    else:
                        self.log(f"Erro de Playwright ao processar o item {nro_cotacao}: {error_message}", "ERRO")
                        # Resetar a página para o próximo item
                        self._reset_session_by_relogging(nro_cotacao, atraso_fases) # Relogin em caso de erro
                except Exception as e_item:
                    self.log(f"Erro inesperado ao processar o item {nro_cotacao}: {e_item}", "ERRO")
                    # Resetar a página para o próximo item
                    self._reset_session_by_relogging(nro_cotacao, atraso_fases) # Relogin em caso de erro
            self.log("--- LOOP DE PROCESSAMENTO CONCLUÍDO ---", "FASE")
            self.status_var.set("Finalizado")

        except Exception as e:
            self.log(f"Erro crítico na automação (Fases 2-5): {e}", "ERRO")
            self.log(f"Traceback: {traceback.format_exc()}", "DEBUG")
            self.status_var.set("Erro! Verifique os logs.")
        finally:
            self.log("Reabilitando botões...", "DEBUG")
            self.btn_iniciar.config(state="normal", text="▶ Iniciar Automação")
            self.btn_pause.config(state="disabled", text="Pausar")
            self.is_paused = False
            self.pause_event.set()


if __name__ == "__main__":
    try:
        if os.path.exists(LOG_FILENAME):
            os.remove(LOG_FILENAME)
            
        root = tk.Tk()
        app = AutomacaoUI(root)
        root.mainloop()
        
    except Exception as e:
        print(f"Erro fatal ao iniciar a aplicação: {e}")
        with open(LOG_FILENAME, "a", encoding="utf-8") as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [FATAL]: {e}\n")
            import traceback
            f.write(f"Traceback: {traceback.format_exc()}\n")
