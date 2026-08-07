import threading
import time
from playwright.sync_api import Page, Error
from typing import Callable, Dict, Optional
import openpyxl
from openpyxl.styles import PatternFill
import os
from ..companies import get_company

excel_lock = threading.Lock()

def registrar_sucesso_em_planilha(
    dados_linha: Dict[str, str],
    log_callback: Callable,
    output_filepath: str,
):
    """
    Encontra a linha correspondente na aba 'Dados Processados', atualiza o status
    para 'Concluído' e colore a linha de verde.
    """
    if not output_filepath:
        nro_cotacao = dados_linha.get('Nro cotação', 'N/A')
        log_callback(f"[F5] Não foi possível registrar o sucesso por falta do arquivo de output.", "ERRO")
        return

    try:
        with excel_lock:
            nro_cotacao = dados_linha.get('Nro cotação', 'N/A')
            workbook = openpyxl.load_workbook(output_filepath)
            
            if "Dados Processados" in workbook.sheetnames:
                ws = workbook["Dados Processados"]
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                headers = [cell.value for cell in ws[1]]
                status_col_index = headers.index("Status") + 1 if "Status" in headers else -1

                for row in ws.iter_rows(min_row=2):
                    if str(row[0].value) == str(nro_cotacao):
                        if status_col_index != -1:
                            row[status_col_index - 1].value = "Concluído"
                        for cell in row:
                            cell.fill = green_fill
                        log_callback(f"[F5] [Item {nro_cotacao}] Status atualizado para 'Concluído' e linha colorida na planilha.", "DEBUG")
            workbook.save(output_filepath)
            log_callback(f"[F5] [Item {nro_cotacao}] Planilha salva com sucesso em: {output_filepath}", "DEBUG")
    except Exception as e:
        log_callback(f"[F5] ERRO CRÍTICO: Falha inesperada ao registrar sucesso na planilha: {e}", "ERRO")

def registrar_erro_em_planilha(
    dados_linha: Dict[str, str],
    motivo: str,
    log_callback: Callable,
    # O parâmetro output_filepath já deve estar aqui.
    # Se o erro "takes 3 positional arguments" ainda ocorre,
    # significa que esta linha não foi atualizada no arquivo.
    output_filepath: str, 
):
    """
    Encontra a planilha de output e registra uma linha na aba 'Contrato não realizado'.
    Registra uma linha na aba 'Contrato não realizado' da planilha de output especificada.
    """
    if not output_filepath:
        nro_cotacao = dados_linha.get('Nro cotação', 'N/A')
        log_callback(f"[F3] [Item {nro_cotacao}] Não foi possível registrar o erro por falta do arquivo de output.", "ERRO")
        return

    try:
        with excel_lock:
            nro_cotacao = dados_linha.get('Nro cotação', 'N/A')
            log_callback(f"[F3] [Item {nro_cotacao}] Registrando erro na planilha: {motivo}", "DEBUG")
            workbook = openpyxl.load_workbook(output_filepath)
            
            if "Contrato não realizado" not in workbook.sheetnames:
                 log_callback(f"[F3] ERRO CRÍTICO: A aba 'Contrato não realizado' não foi encontrada na planilha '{os.path.basename(output_filepath)}'.", "ERRO")
                 return

            # --- LÓGICA PARA REMOVER A LINHA DA ABA 'DADOS PROCESSADOS' ---
            if "Dados Processados" in workbook.sheetnames:
                ws_processados = workbook["Dados Processados"]
                linha_para_deletar = -1
                # Itera de baixo para cima para evitar problemas de índice ao deletar
                for i in range(ws_processados.max_row, 1, -1):
                    # A coluna 'A' (índice 1) contém o 'Nro cotação'
                    if str(ws_processados.cell(row=i, column=1).value) == str(nro_cotacao):
                        linha_para_deletar = i
                        break
                
                if linha_para_deletar != -1:
                    ws_processados.delete_rows(linha_para_deletar)
                    log_callback(f"[F3] [Item {nro_cotacao}] Linha correspondente removida da aba 'Dados Processados'.", "DEBUG")
                else:
                    log_callback(f"[F3] [Item {nro_cotacao}] AVISO: Não foi possível encontrar a linha para remover da aba 'Dados Processados'.", "AVISO")

            sheet = workbook["Contrato não realizado"]
            
            linha_erro = [
                dados_linha.get("Nro cotação", ""),
                dados_linha.get("Categoria veículo", ""),
                dados_linha.get("Cidade", ""),
                dados_linha.get("UF", ""),
                dados_linha.get("Nome", ""),
                dados_linha.get("Placa", ""),
                dados_linha.get("Data pagamento", ""),
                motivo,
                "Erro" # Status
            ]
            
            sheet.append(linha_erro)
            workbook.save(output_filepath)
            log_callback(f"[F3] [Item {nro_cotacao}] Planilha de erros salva com sucesso em: {output_filepath}", "DEBUG")
            log_callback(f"[F3] [Item {nro_cotacao}] Erro registrado com sucesso na planilha.", "INFO")

    except FileNotFoundError:
        log_callback(f"[F3] ERRO CRÍTICO: Não foi possível encontrar a planilha de output em '{output_filepath}' para registrar o erro.", "ERRO")
    except Exception as e:
        log_callback(f"[F3] ERRO CRÍTICO: Falha inesperada ao registrar erro na planilha: {e}", "ERRO")
        import traceback
        log_callback(f"[F3] Traceback (Planilha): {traceback.format_exc()}", "DEBUG")

def preencher_formulario(
    page: Page, 
    dados_linha: Dict[str, str], 
    log_callback: Callable[[str, str], None],
    pause_event: threading.Event, 
    output_filepath: str,
    atraso_etapas: float,
    atraso_fases: float
) -> bool:
    """
    Executa os passos de preenchimento da Fase 3.
    """
    try:
        nro_cotacao_item = dados_linha.get("Nro cotação", "N/A")
        log_callback(f"[F3] [Item {nro_cotacao_item}] --- INÍCIO: PREENCHIMENTO BÁSICO ---", "FASE")
        
        nro_cotacao = nro_cotacao_item

        # ETAPA 1: Selecionar "LOGTUDO MATRIZ - BAHIA"
        pause_event.wait()
        log_callback(f"[F3] [Item {nro_cotacao_item}] Etapa 1: Selecionando Agência...", "DEBUG")
        page.select_option('select[name="dados_agencias_id"]', value="2")
        time.sleep(atraso_etapas)

        # ETAPA 2: Selecionar "Encarte Bahia"
        pause_event.wait()
        log_callback(f"[F3] [Item {nro_cotacao_item}] Etapa 2: Selecionando Talão...", "DEBUG")
        page.select_option('select[name="dados_tiposTaloes_id"]', value="53")
        time.sleep(atraso_etapas)

        # ETAPA 3: Selecionar checkbox "emitirReciboFrete"
        pause_event.wait()
        log_callback(f"[F3] [Item {nro_cotacao_item}] Etapa 3: Marcando Checkbox 'Emitir Recibo'...", "DEBUG")
        page.check('input[name="dados_emitirReciboFrete[]"]')
        time.sleep(atraso_etapas)

        # ETAPA 4-5.1: Executar pesquisa e seleção da cotação
        pause_event.wait()
        remetente_str = str(dados_linha.get("Remetente", ""))
        company = get_company(remetente_str)
        
        ok_pesquisa = company.executar_pesquisa_cotacao(
            page=page,
            dados_linha=dados_linha,
            nro_cotacao=nro_cotacao,
            log_callback=log_callback,
            atraso_etapas=atraso_etapas,
            output_filepath=output_filepath
        )
        if not ok_pesquisa:
            return False

        # ETAPA 6: Substituir valor em "dados_complementoPedido"
        pause_event.wait()
        valor_pedido = company.get_complemento_pedido(nro_cotacao_item)
        if valor_pedido is not None:
            log_callback(f"[F3] [Item {nro_cotacao_item}] Etapa 6: Preenchendo Complemento do Pedido com '{valor_pedido}'...", "DEBUG")
            page.fill('input[name="dados_complementoPedido"]', valor_pedido, force=True)
            time.sleep(atraso_etapas)

        # ETAPA 7: Clicar em "Avançar >>"
        pause_event.wait()
        log_callback(f"[F3] [Item {nro_cotacao_item}] Etapa 7: Clicando em 'Avançar'...", "DEBUG")
        
        botao_avancar_selector = '#botao_avancar'
        primeiro_campo_fase4_selector = 'input[name="pesquisa_enderecoDestinatario_id"]'

        page.click(botao_avancar_selector)
        
        log_callback(f"[F3] [Item {nro_cotacao_item}] Aguardando transição para a Fase 4...", "DEBUG")
        page.wait_for_selector(primeiro_campo_fase4_selector, state="visible", timeout=30000)

        log_callback(f"[F3] [Item {nro_cotacao_item}] Fase concluída com sucesso.", "SUCESSO")
        return True

    except Error as e:
        log_callback(f"[F3] Erro de Playwright na Fase 3 (Item {nro_cotacao_item}): {e}", "ERRO")
        return False
    except Exception as e:
        log_callback(f"[F3] Erro inesperado na Fase 3 (Item {nro_cotacao_item}): {e}", "ERRO")
        import traceback
        log_callback(f"[F3] Traceback Fase 3: {traceback.format_exc()}", "DEBUG")
        return False