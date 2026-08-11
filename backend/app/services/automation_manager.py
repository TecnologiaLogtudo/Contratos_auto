from __future__ import annotations

import configparser
import json
import os
import re
import shutil
import threading
import time
import traceback
import sqlite3
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import xlrd

from ..phases import (
    fase1_processamento,
    fase2_login,
    fase3_preenchimento,
    fase4_frete,
    fase5_contrato_frete,
)

from ..schemas import AutomationConfig, JobStatus, LogEvent

BASE_DIR = Path(__file__).resolve().parents[3]
DATA_DIR = Path(__file__).resolve().parents[1] / "data"
UPLOADS_DIR = DATA_DIR / "uploads"
RESULTS_DIR = DATA_DIR / "results"
ARTIFACTS_DIR = DATA_DIR / "artifacts"
LOGS_HISTORY_PATH = DATA_DIR / "logs_history.json"
CONFIG_PATH = BASE_DIR / "config.ini"
RESULTS_HISTORY_PATH = DATA_DIR / "results_history.json"
DB_PATH = DATA_DIR / "automation.db"
URL_DESTINO = "https://logtudo.e-login.net/versoes/versao5.0/rotinas/formulario.php?rotina=trans_conhecimento&OP=O1&_qsf=1"

PHASE_RE = re.compile(r"\[(F\d)\]")
COTACAO_RE = re.compile(r"Cotação:\s*([^\]\s]+)")


def _resolve_artifact_disk_path(file_path: Optional[str], job_id: Optional[str] = None) -> Optional[Path]:
    """
    Resolve o caminho físico de um arquivo de artefato no disco rígido.
    Suporta caminhos absolutos, caminhos relativos ao workspace e caminhos legado
    do contêiner (como /app/backend/app/data/... ou /app/webapp/exports/...)
    """
    if not file_path:
        return None
    clean_path = str(file_path).replace("\\", "/")
    path_direct = Path(file_path)
    if path_direct.exists() and path_direct.is_file():
        return path_direct.resolve()
    filename = path_direct.name
    candidates = [
        DATA_DIR / filename,
        ARTIFACTS_DIR / filename,
        RESULTS_DIR / filename,
        UPLOADS_DIR / filename,
    ]
    if job_id:
        candidates.append(ARTIFACTS_DIR / job_id / filename)
        candidates.append(DATA_DIR / "exports" / "jobs" / job_id / filename)
    if "/backend/app/data/" in clean_path:
        parts = clean_path.split("/backend/app/data/")
        if len(parts) > 1:
            candidates.append(DATA_DIR / parts[1])
    if "exports/jobs" in clean_path:
        parts = clean_path.split("exports/jobs/")
        if len(parts) > 1:
            sub_parts = parts[1].split("/")
            if len(sub_parts) > 0:
                candidates.append(ARTIFACTS_DIR / sub_parts[-1])
                if job_id:
                    candidates.append(ARTIFACTS_DIR / job_id / sub_parts[-1])
    if "uploads/" in clean_path:
        parts = clean_path.split("uploads/")
        if len(parts) > 1:
            candidates.append(UPLOADS_DIR / parts[-1])
    candidates.append(Path("backend/app/data/artifacts") / filename)
    candidates.append(Path("backend/app/data/results") / filename)
    candidates.append(Path("backend/app/data/uploads") / filename)
    if job_id:
        candidates.append(Path("backend/app/data/artifacts") / job_id / filename)
    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_file():
                return candidate.resolve()
        except Exception:
            continue
    return None


def init_db() -> None:
    """
    Inicializa o banco de dados SQLite e cria as tabelas necessárias,
    incluindo a tabela job_artifacts para gerenciar metadados de arquivos.
    """
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS results_history (
                id TEXT PRIMARY KEY,
                job_id TEXT,
                log_session_id TEXT,
                arquivo_original TEXT,
                result_file TEXT,
                status TEXT,
                total INTEGER,
                processados INTEGER,
                sucessos INTEGER,
                erros INTEGER,
                pendentes INTEGER,
                taxa_sucesso REAL,
                created_at TEXT
            )
        """)
        cursor.execute("PRAGMA table_info(results_history)")
        existing_columns = {row[1] for row in cursor.fetchall()}
        if "pendentes" not in existing_columns:
            cursor.execute("ALTER TABLE results_history ADD COLUMN pendentes INTEGER DEFAULT 0")
        if "taxa_sucesso" not in existing_columns:
            cursor.execute("ALTER TABLE results_history ADD COLUMN taxa_sucesso REAL DEFAULT 0.0")
        if "duracao" not in existing_columns:
            cursor.execute("ALTER TABLE results_history ADD COLUMN duracao TEXT DEFAULT ''")
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS logs_history (
                id TEXT PRIMARY KEY,
                job_id TEXT,
                created_at TEXT,
                status TEXT,
                user TEXT,
                ip TEXT,
                acoes_criticas TEXT,
                passos_acoes TEXT,
                artefatos TEXT,
                browser_logs TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS job_artifacts (
                id TEXT PRIMARY KEY,
                job_id TEXT,
                type TEXT,
                file_path TEXT,
                created_at TEXT
            )
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[DB] Erro ao inicializar banco de dados SQLite: {e}")


class StopRequestedException(BaseException):
    pass


class JobPauseEvent:
    def __init__(self, stop_event: threading.Event):
        self._event = threading.Event()
        self._stop_event = stop_event

    def wait(self, timeout: Optional[float] = None) -> bool:
        if self._stop_event.is_set():
            raise StopRequestedException("Parada solicitada pelo usuário.")
        while not self._event.is_set():
            if self._stop_event.is_set():
                raise StopRequestedException("Parada solicitada pelo usuário.")
            self._event.wait(timeout=0.1)
        if self._stop_event.is_set():
            raise StopRequestedException("Parada solicitada pelo usuário.")
        return True

    def is_set(self) -> bool:
        return self._event.is_set()

    def set(self) -> None:
        self._event.set()

    def clear(self) -> None:
        self._event.clear()


@dataclass
class JobRuntime:
    status: JobStatus
    config: AutomationConfig
    source_file: Path
    stored_file: Path
    logs: list[LogEvent] = field(default_factory=list)
    seq: int = 0
    lock: threading.Lock = field(default_factory=threading.Lock)
    condition: threading.Condition = field(default_factory=lambda: threading.Condition(threading.Lock()))
    pause_event: threading.Event = field(default_factory=threading.Event)
    stop_event: threading.Event = field(default_factory=threading.Event)
    thread: Optional[threading.Thread] = None

    playwright_instance: object = None
    browser: object = None
    browser_context: object = None
    page: object = None
    planilha_processada_path: Optional[str] = None
    log_session_id: str = ""
    requester_ip: str = ""
    configured_user: str = ""
    critical_actions: list[dict] = field(default_factory=list)
    browser_logs: list[dict] = field(default_factory=list)
    artifacts: list[dict] = field(default_factory=list)
    start_time: float = field(default_factory=time.time)

    def __post_init__(self) -> None:
        self.pause_event = JobPauseEvent(self.stop_event)

    def emit(self, message: str, level: str = "INFO") -> None:
        with self.lock:
            self.seq += 1
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            phase = None
            cotacao = None
            phase_m = PHASE_RE.search(message)
            cotacao_m = COTACAO_RE.search(message)
            if phase_m:
                phase = phase_m.group(1)
                self.status.current_phase = phase
            if cotacao_m:
                cotacao = cotacao_m.group(1)
                self.status.current_cotacao = cotacao

            event = LogEvent(
                seq=self.seq,
                timestamp=ts,
                level=level.upper(),
                message=message,
                phase=phase,
                cotacao=cotacao,
            )
            self.logs.append(event)
            step = event.model_dump()
            if level.upper() in {"ERRO", "AVISO"}:
                self.critical_actions.append(step)
        with self.condition:
            self.condition.notify_all()

    def take_screenshot(self, name: str) -> None:
        """
        Captura uma captura de tela (screenshot) da página atual do Playwright
        e a registra como artefato no banco de dados.
        """
        try:
            if not self.page:
                return
            clean_name = "".join(c for c in name if c.isalnum() or c in (" ", "_", "-")).rstrip()
            clean_name = clean_name.replace(" ", "_")
            filename = f"{self.log_session_id}_{clean_name}_{int(time.time())}.png"
            filepath = ARTIFACTS_DIR / filename
            
            # Captura a screenshot usando Playwright
            self.page.screenshot(path=str(filepath))
            
            with self.lock:
                self.artifacts.append({
                    "type": "printscreen",
                    "name": filename,
                    "path": str(filepath.resolve())
                })
                
            # Registra o artefato no banco de dados SQLite de forma assíncrona/segura
            try:
                manager._record_artifact(self.status.id, "printscreen", filepath)
            except Exception as db_err:
                self.emit(f"[DB] Falha ao registrar screenshot no banco de dados: {db_err}", "DEBUG")
                
            self.emit(f"Printscreen salvo: {filename}", "DEBUG")
        except Exception as e:
            self.emit(f"Falha ao tirar printscreen ({name}): {e}", "DEBUG")



class AutomationManager:
    def __init__(self) -> None:
        UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
        RESULTS_DIR.mkdir(parents=True, exist_ok=True)
        ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True)
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        self._jobs: dict[str, JobRuntime] = {}
        self._jobs_lock = threading.Lock()
        self._history_lock = threading.RLock()
        
        # Inicializa banco SQLite
        init_db()

        if not RESULTS_HISTORY_PATH.exists():
            RESULTS_HISTORY_PATH.write_text("[]", encoding="utf-8")
        if not LOGS_HISTORY_PATH.exists():
            LOGS_HISTORY_PATH.write_text("[]", encoding="utf-8")

    def _record_artifact(self, job_id: str, type_str: str, file_path: Path) -> None:
        """
        Registra o metadado de um artefato na tabela job_artifacts do banco de dados SQLite.
        """
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            artifact_id = str(uuid.uuid4())
            cursor.execute("""
                INSERT INTO job_artifacts (id, job_id, type, file_path, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (
                artifact_id,
                job_id,
                type_str,
                str(file_path.resolve()),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"[DB] Erro ao gravar metadado do artefato {type_str}: {e}")

    def load_config(self) -> AutomationConfig:
        parser = configparser.ConfigParser()
        if CONFIG_PATH.exists():
            parser.read(CONFIG_PATH, encoding="utf-8")

        # Prioridade: Variáveis de Ambiente -> config.ini -> fallback padrão
        return AutomationConfig(
            login=os.getenv("AUTOMACAO_LOGIN", parser.get("CREDENCIAS", "login", fallback="")),
            senha=os.getenv("AUTOMACAO_SENHA", parser.get("CREDENCIAS", "senha", fallback="")),
            atraso_fases=float(os.getenv("AUTOMACAO_ATRASOFASES", parser.getfloat("AUTOMACAO", "atrasofases", fallback=0.5))),
            atraso_etapas=float(os.getenv("AUTOMACAO_ATRASOETAPAS", parser.getfloat("AUTOMACAO", "atrasoetapas", fallback=0.15))),
            dados_km=os.getenv("AUTOMACAO_DADOS_KM", parser.get("AUTOMACAO", "dados_km", fallback="20")),
            aceitar_frete_minimo_antt=os.getenv("AUTOMACAO_ACEITAR_FRETE_MINIMO", str(parser.getboolean("AUTOMACAO", "aceitarfreteminimoantt", fallback=True))).lower() in ("true", "1", "yes"),
        )

    def save_config(self, config: AutomationConfig) -> AutomationConfig:
        parser = configparser.ConfigParser()
        parser["CREDENCIAS"] = {
            "login": config.login,
            "senha": config.senha,
        }
        parser["AUTOMACAO"] = {
            "atrasofases": str(config.atraso_fases),
            "atrasoetapas": str(config.atraso_etapas),
            "dados_km": config.dados_km,
            "aceitarfreteminimoantt": str(config.aceitar_frete_minimo_antt),
        }
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            parser.write(f)
        return config

    def preview_file(self, file_path: Path) -> dict:
        suffix = file_path.suffix.lower()
        if suffix == ".xls":
            return self._preview_xls(file_path)
        return self._preview_xlsx(file_path)

    def _preview_xlsx(self, file_path: Path) -> dict:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            headers_raw = [c.value for c in ws[1]] if ws.max_row > 0 else []
            headers = [str(h) if h is not None else "" for h in headers_raw]
            preview = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                if i > 10:
                    break
                row_dict = {}
                for idx, value in enumerate(row):
                    key = headers[idx] if idx < len(headers) and headers[idx] else f"col_{idx+1}"
                    row_dict[key] = value
                preview.append(row_dict)
            rows = max(0, ws.max_row - 1)
            cols = ws.max_column
            return {"filename": file_path.name, "rows": rows, "cols": cols, "headers": headers, "preview": preview}
        finally:
            wb.close()

    def _preview_xls(self, file_path: Path) -> dict:
        wb = xlrd.open_workbook(str(file_path), formatting_info=False)
        sheet = wb.sheet_by_index(0)
        headers_raw = sheet.row_values(0) if sheet.nrows > 0 else []
        headers = [str(h) if h is not None else "" for h in headers_raw]
        preview = []
        max_preview = min(sheet.nrows, 11)
        for row_idx in range(1, max_preview):
            row = sheet.row_values(row_idx)
            row_dict = {}
            for idx, value in enumerate(row):
                key = headers[idx] if idx < len(headers) and headers[idx] else f"col_{idx+1}"
                row_dict[key] = value
            preview.append(row_dict)
        rows = max(0, sheet.nrows - 1)
        cols = sheet.ncols
        return {"filename": file_path.name, "rows": rows, "cols": cols, "headers": headers, "preview": preview}

    def create_job(self, source_file: Path, config: AutomationConfig, requester_ip: str = "") -> JobRuntime:
        """
        Cria um novo Job de automação, copia a planilha de entrada e registra a mesma
        como o primeiro artefato (input_file).
        """
        job_id = str(uuid.uuid4())
        # log_session_id agora inclui segundos e um identificador único para garantir unicidade real
        log_session_id = f"{datetime.now().strftime('%d%m%Y-%H%M%S')}-{uuid.uuid4().hex[:6]}"
        stored_file = UPLOADS_DIR / f"{job_id}_{source_file.name}"
        shutil.copyfile(source_file, stored_file)

        runtime = JobRuntime(
            status=JobStatus(id=job_id, log_session_id=log_session_id, state="running", message="Inicializando"),
            config=config,
            source_file=source_file,
            stored_file=stored_file,
            log_session_id=log_session_id,
            requester_ip=requester_ip,
            configured_user=(config.login or "").strip(),
        )
        runtime.pause_event.set()
        runtime.thread = threading.Thread(target=self._run_job, args=(runtime,), daemon=True)
        with self._jobs_lock:
            self._jobs[job_id] = runtime
            
        # Registra a planilha de entrada (input_file) no banco de dados de artefatos
        self._record_artifact(job_id, "input_file", stored_file)
        
        runtime.thread.start()
        return runtime

    def get_job(self, job_id: str) -> Optional[JobRuntime]:
        with self._jobs_lock:
            return self._jobs.get(job_id)

    def pause(self, job_id: str) -> JobStatus:
        job = self._must_get(job_id)
        job.pause_event.clear()
        with job.lock:
            if job.status.state == "running":
                job.status.state = "paused"
                job.status.message = "Pausado"
        job.emit("Execução pausada pelo usuário.", "AVISO")
        return job.status

    def resume(self, job_id: str) -> JobStatus:
        job = self._must_get(job_id)
        job.pause_event.set()
        with job.lock:
            if job.status.state == "paused":
                job.status.state = "running"
                job.status.message = "Executando"
        job.emit("Execução retomada pelo usuário.", "INFO")
        return job.status

    def stop(self, job_id: str) -> JobStatus:
        job = self._must_get(job_id)
        job.stop_event.set()
        job.pause_event.set()
        with job.lock:
            if job.status.state in {"running", "paused"}:
                job.status.state = "stopped"
                job.status.message = "Parado pelo usuário"
        job.emit("Parada solicitada pelo usuário.", "AVISO")
        return job.status


    def list_results_history(self) -> list[dict]:
        # Tenta ler do SQLite primeiro
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM results_history ORDER BY created_at DESC")
            rows = cursor.fetchall()
            conn.close()
            return [dict(r) for r in rows]
        except Exception as e:
            print(f"[DB] Erro ao ler histórico de resultados do SQLite: {e}")

        # Fallback para JSON
        return self._load_results_history_json()

    def _load_results_history_json(self) -> list[dict]:
        with self._history_lock:
            try:
                data = json.loads(RESULTS_HISTORY_PATH.read_text(encoding="utf-8"))
                return sorted(data, key=lambda x: x.get("created_at", ""), reverse=True)
            except Exception:
                return []

    def delete_results(self, ids: list[str], password: str) -> dict:
        self._validate_admin_reset_password(password)

        # Exclui do SQLite primeiro
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            placeholders = ",".join(["?"] * len(ids))
            cursor.execute(f"SELECT result_file FROM results_history WHERE id IN ({placeholders})", ids)
            rows = cursor.fetchall()
            for r in rows:
                if r[0]:
                    Path(r[0]).unlink(missing_ok=True)
            
            cursor.execute(f"DELETE FROM results_history WHERE id IN ({placeholders})", ids)
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"[DB] Erro ao excluir resultados do SQLite: {e}")

        # Fallback para JSON
        with self._history_lock:
            history = self._load_results_history_json()
            keep = []
            removed = 0
            for item in history:
                if item.get("id") in ids:
                    path = item.get("result_file")
                    if path:
                        Path(path).unlink(missing_ok=True)
                    removed += 1
                else:
                    keep.append(item)
            try:
                RESULTS_HISTORY_PATH.write_text(json.dumps(keep, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception as e:
                print(f"[JSON] Erro ao salvar fallback JSON de exclusão: {e}")

        return {"removed": len(ids)}

    def _record_artifact_once(self, job_id: str, type_str: str, file_path: Path) -> None:
        conn = None
        exists = None
        try:
            resolved_path = str(file_path.resolve())
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id FROM job_artifacts WHERE job_id = ? AND type = ? AND file_path = ? LIMIT 1",
                (job_id, type_str, resolved_path),
            )
            exists = cursor.fetchone()
        except Exception as e:
            print(f"[DB] Erro ao verificar artefato {type_str}: {e}")
        finally:
            if conn:
                conn.close()
        if not exists:
            self._record_artifact(job_id, type_str, file_path)

    def _finalize_result_file(self, job: JobRuntime) -> None:
        if not job.planilha_processada_path:
            return

        source_res = Path(job.planilha_processada_path)
        if not source_res.exists() or not source_res.is_file():
            return

        final_res = RESULTS_DIR / f"{job.status.id}_{source_res.name}"
        try:
            if source_res.resolve() != final_res.resolve():
                shutil.copyfile(source_res, final_res)
            else:
                final_res = source_res
        except Exception as e:
            job.emit(f"Erro ao copiar planilha final para histórico: {e}", "ERRO")
            return

        with job.lock:
            job.status.result_file = str(final_res.resolve())
        self._record_artifact_once(job.status.id, "result_file", final_res)

    @staticmethod
    def _normalize_status(value: object) -> str:
        return str(value or "").strip().lower().replace("í", "i")

    def _count_result_statuses(self, result_file: Optional[str]) -> dict:
        counts = {"sucessos": 0, "erros": 0, "pendentes": 0}
        if not result_file:
            return counts

        path = Path(result_file)
        if not path.exists() or not path.is_file():
            return counts

        try:
            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            try:
                if "Dados Processados" in workbook.sheetnames:
                    ws = workbook["Dados Processados"]
                    headers = [cell.value for cell in ws[1]]
                    status_col = headers.index("Status") + 1 if "Status" in headers else None
                    if status_col:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            status = self._normalize_status(row[status_col - 1] if len(row) >= status_col else "")
                            if status == "concluido":
                                counts["sucessos"] += 1
                            elif status == "pendente":
                                counts["pendentes"] += 1

                if "Contrato não realizado" in workbook.sheetnames:
                    ws = workbook["Contrato não realizado"]
                    headers = [cell.value for cell in ws[1]]
                    status_col = headers.index("Status") + 1 if "Status" in headers else None
                    if status_col:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            status = self._normalize_status(row[status_col - 1] if len(row) >= status_col else "")
                            if status == "erro":
                                counts["erros"] += 1
            finally:
                workbook.close()
        except Exception as e:
            print(f"[RESULTS] Erro ao contabilizar planilha final: {e}")

        return counts

    def _append_result_history(self, job: JobRuntime) -> None:
        counts = self._count_result_statuses(job.status.result_file)
        total_contabilizado = counts["sucessos"] + counts["erros"] + counts["pendentes"]
        taxa_sucesso = round((counts["sucessos"] / total_contabilizado) * 100, 2) if total_contabilizado else 0.0
        
        duration_s = time.time() - job.start_time
        mins = int(duration_s // 60)
        secs = int(duration_s % 60)
        duracao_str = f"{mins}m {secs}s"
        
        item = {
            "id": job.status.id,
            "job_id": job.status.id,
            "log_session_id": job.log_session_id,
            "arquivo_original": job.stored_file.name,
            "result_file": job.status.result_file,
            "status": job.status.state,
            "total": total_contabilizado,
            "processados": counts["sucessos"] + counts["erros"],
            "sucessos": counts["sucessos"],
            "erros": counts["erros"],
            "pendentes": counts["pendentes"],
            "taxa_sucesso": taxa_sucesso,
            "duracao": duracao_str,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        # Salva no SQLite
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                INSERT OR REPLACE INTO results_history (
                    id, job_id, log_session_id, arquivo_original, result_file, status, total, processados, sucessos, erros, pendentes, taxa_sucesso, duracao, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                item["id"], item["job_id"], item["log_session_id"], item["arquivo_original"], item["result_file"],
                item["status"], item["total"], item["processados"], item["sucessos"], item["erros"],
                item["pendentes"], item["taxa_sucesso"], item["duracao"], item["created_at"]
            ))
            conn.commit()
            conn.close()
        except Exception as e:
            job.emit(f"[DB] Erro ao salvar histórico de resultados no SQLite: {e}", "ERRO")

        # Fallback JSON
        with self._history_lock:
            try:
                history = self._load_results_history_json()
                history = [h for h in history if h.get("id") != item["id"] and h.get("job_id") != item["job_id"]]
                history.append(item)
                RESULTS_HISTORY_PATH.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception as e:
                job.emit(f"[JSON] Erro ao salvar histórico de resultados no fallback: {e}", "ERRO")

    def _append_logs_history(self, job: JobRuntime) -> None:
        item = {
            "id": job.log_session_id,
            "job_id": job.status.id,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status": job.status.state,
            "user": job.configured_user,
            "ip": job.requester_ip,
            "acoes_criticas": job.critical_actions,
            "passos_acoes": [e.model_dump() for e in job.logs],
            "artefatos": job.artifacts,
            "browser_logs": job.browser_logs,
        }

        # Salva no SQLite
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                INSERT OR REPLACE INTO logs_history (
                    id, job_id, created_at, status, user, ip, acoes_criticas, passos_acoes, artefatos, browser_logs
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                item["id"], item["job_id"], item["created_at"], item["status"], item["user"], item["ip"],
                json.dumps(item["acoes_criticas"], ensure_ascii=False),
                json.dumps(item["passos_acoes"], ensure_ascii=False),
                json.dumps(item["artefatos"], ensure_ascii=False),
                json.dumps(item["browser_logs"], ensure_ascii=False)
            ))
            conn.commit()
            conn.close()
        except Exception as e:
            job.emit(f"[DB] Erro ao salvar log de sessão no SQLite: {e}", "ERRO")

        # Fallback JSON
        with self._history_lock:
            try:
                sessions = self._load_logs_history_json()
                sessions = [s for s in sessions if s.get("id") != item["id"]]
                sessions.append(item)
                LOGS_HISTORY_PATH.write_text(json.dumps(sessions, ensure_ascii=False, indent=2), encoding="utf-8")
            except Exception as e:
                job.emit(f"[JSON] Erro ao salvar log de sessão no fallback: {e}", "ERRO")

    def list_log_sessions(self) -> list[dict]:
        # Tenta ler do SQLite primeiro
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM logs_history ORDER BY created_at DESC")
            rows = cursor.fetchall()
            conn.close()
            sessions = []
            for r in rows:
                sessions.append({
                    "id": r["id"],
                    "job_id": r["job_id"],
                    "created_at": r["created_at"],
                    "status": r["status"],
                    "user": r["user"],
                    "ip": r["ip"],
                    "acoes_criticas": json.loads(r["acoes_criticas"] or "[]"),
                    "passos_acoes": json.loads(r["passos_acoes"] or "[]"),
                    "artefatos": json.loads(r["artefatos"] or "[]"),
                    "browser_logs": json.loads(r["browser_logs"] or "[]")
                })
            return sessions
        except Exception as e:
            print(f"[DB] Erro ao ler logs de sessões do SQLite: {e}")

        # Fallback para JSON
        return self._load_logs_history_json()

    def _load_logs_history_json(self) -> list[dict]:
        with self._history_lock:
            try:
                data = json.loads(LOGS_HISTORY_PATH.read_text(encoding="utf-8"))
                return sorted(data, key=lambda x: x.get("created_at", ""), reverse=True)
            except Exception:
                return []

    def get_log_session(self, session_id: str) -> Optional[dict]:
        # Tenta SQLite primeiro
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM logs_history WHERE id = ?", (session_id,))
            r = cursor.fetchone()
            conn.close()
            if r:
                return {
                    "id": r["id"],
                    "job_id": r["job_id"],
                    "created_at": r["created_at"],
                    "status": r["status"],
                    "user": r["user"],
                    "ip": r["ip"],
                    "acoes_criticas": json.loads(r["acoes_criticas"] or "[]"),
                    "passos_acoes": json.loads(r["passos_acoes"] or "[]"),
                    "artefatos": json.loads(r["artefatos"] or "[]"),
                    "browser_logs": json.loads(r["browser_logs"] or "[]")
                }
        except Exception as e:
            print(f"[DB] Erro ao ler log de sessão individual do SQLite: {e}")

        # Fallback para JSON
        sessions = self._load_logs_history_json()
        return next((s for s in sessions if s.get("id") == session_id), None)

    def clear_logs(self, password: str) -> dict:
        self._validate_admin_reset_password(password)

        # Limpa do SQLite
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM logs_history")
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"[DB] Erro ao limpar logs do SQLite: {e}")

        # Fallback JSON
        with self._history_lock:
            try:
                LOGS_HISTORY_PATH.write_text("[]", encoding="utf-8")
            except Exception as e:
                print(f"[JSON] Erro ao limpar logs no fallback JSON: {e}")

        return {"cleared": True}


    def _validate_admin_reset_password(self, password: str) -> None:
        admin_password = os.getenv("ADMIN_RESET_PASSWORD", "").strip()
        if not admin_password:
            raise PermissionError("ADMIN_RESET_PASSWORD não configurada no ambiente.")
        if password != admin_password:
            raise PermissionError("Senha inválida")

    def _must_get(self, job_id: str) -> JobRuntime:
        job = self.get_job(job_id)
        if not job:
            raise KeyError(job_id)
        return job

    def _update_progress(self, job: JobRuntime, current: int, total: int, message: str = "") -> None:
        with job.lock:
            job.status.current_item = current
            job.status.total_items = total
            job.status.percent = round((current / total) * 100, 2) if total else 0
            if message:
                job.status.message = message

    def _run_job(self, job: JobRuntime) -> None:
        cfg = job.config
        try:
            if not cfg.login or not cfg.senha:
                raise ValueError("Login e senha são obrigatórios para iniciar a automação.")

            job.emit("[F0] Job iniciado.", "INFO")
            with job.lock:
                job.status.state = "running"
                job.status.message = "Processando planilha"

            filepath = str(job.stored_file)
            if self._is_planilha_tratada(filepath):
                job.planilha_processada_path = filepath
                job.emit("[F1] Planilha já tratada detectada. Pulando Fase 1.", "INFO")
            else:
                job.emit("[F1] Iniciando processamento da planilha...", "INFO")
                job.planilha_processada_path = fase1_processamento.processar_planilha(filepath, job.emit)
                if not job.planilha_processada_path:
                    raise RuntimeError("Falha ao processar planilha na Fase 1.")

            with job.lock:
                job.status.message = "Login e setup"

            job.playwright_instance, job.browser, job.browser_context = fase2_login.launch_browser(job.emit)
            if not job.browser or not job.browser_context:
                raise RuntimeError("Falha ao iniciar navegador.")

            # Habilitar Playwright Tracing para depuração completa
            try:
                job.browser_context.tracing.start(screenshots=True, snapshots=True, sources=True)
                job.emit("[F2] Playwright tracing iniciado com sucesso.", "DEBUG")
            except Exception as trace_err:
                job.emit(f"[F2] Aviso: Não foi possível iniciar o Playwright tracing: {trace_err}", "DEBUG")

            job.page = fase2_login.perform_login(
                job.browser_context,
                cfg.login,
                cfg.senha,
                URL_DESTINO,
                job.emit,
                browser_log_callback=lambda m, l="ERROR": job.browser_logs.append(
                    {"timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "level": l, "message": m}
                ),
            )
            if not job.page:
                raise RuntimeError("Falha no login.")
            
            job.take_screenshot("login_sucesso")


            dados_planilha = self._ler_dados_planilha(job)
            dados_para_processar = [item for item in dados_planilha if item.get("Status") == "Pendente"]
            if not dados_para_processar:
                job.emit("Nenhum item pendente encontrado.", "AVISO")
                with job.lock:
                    job.status.state = "completed"
                    job.status.message = "Finalizado sem pendências"
                return

            total = len(dados_para_processar)
            self._update_progress(job, 0, total, "Executando fases 3-5")

            for idx, item in enumerate(dados_para_processar, start=1):
                if job.stop_event.is_set():
                    break
                while not job.pause_event.is_set():
                    if job.stop_event.is_set():
                        break
                    time.sleep(0.2)
                if job.stop_event.is_set():
                    break

                nro = item.get("Nro cotação", "N/A")
                self._update_progress(job, idx - 1, total, f"Processando cotação {nro}")
                job.emit(f"--- Processando Item {idx}/{total} (Cotação: {nro}) ---", "INFO")

                # Determina a estratégia da empresa para direcionar a navegação inicial
                remetente_planilha = item.get("Remetente", "")
                from ..companies import get_company
                company = get_company(remetente_planilha)

                if hasattr(company, "preparar_dados_cotacao"):
                    ok_prep = company.preparar_dados_cotacao(job.page, item, job.emit, cfg.atraso_etapas)
                    if not ok_prep:
                        job.take_screenshot(f"falha_preparacao_cotacao_{nro}")
                        self._sleep_or_stop(job, cfg.atraso_fases)
                        self._reset_session(job, nro, cfg.atraso_fases)
                        continue
                else:
                    # Regras padrão: navega direto para o formulário de Conhecimento
                    try:
                        job.page.goto(URL_DESTINO, wait_until="load", timeout=60000)
                    except Exception as goto_err:
                        job.emit(f"Erro ao navegar para o formulário inicial: {goto_err}", "ERRO")
                        self._reset_session(job, nro, cfg.atraso_fases)
                        continue

                ok3 = fase3_preenchimento.preencher_formulario(job.page, item, job.emit, job.pause_event, job.planilha_processada_path, cfg.atraso_etapas, cfg.atraso_fases)
                if not ok3:
                    job.take_screenshot(f"falha_formulario_cotacao_{nro}")
                    self._sleep_or_stop(job, cfg.atraso_fases)
                    self._reset_session(job, nro, cfg.atraso_fases)
                    continue

                self._sleep_or_stop(job, cfg.atraso_fases)
                ok4 = fase4_frete.preencher_frete(job.page, item, job.emit, job.pause_event, job.planilha_processada_path, cfg.atraso_etapas)
                if not ok4:
                    job.take_screenshot(f"falha_frete_cotacao_{nro}")
                    self._sleep_or_stop(job, cfg.atraso_fases)
                    self._reset_session(job, nro, cfg.atraso_fases)
                    continue

                self._sleep_or_stop(job, cfg.atraso_fases)
                ok5 = fase5_contrato_frete.preencher_contrato_frete(job.page, item, job.emit, job.pause_event, cfg.atraso_etapas, job.planilha_processada_path, cfg.atraso_fases, cfg.dados_km)
                if not ok5:
                    job.take_screenshot(f"falha_contrato_cotacao_{nro}")
                    self._sleep_or_stop(job, cfg.atraso_fases)
                    self._reset_session(job, nro, cfg.atraso_fases)
                    continue

                job.take_screenshot(f"sucesso_contrato_cotacao_{nro}")
                self._sleep_or_stop(job, cfg.atraso_fases)
                self._update_progress(job, idx, total, f"Cotação {nro} concluída")

            with job.lock:
                if job.status.state != "stopped":
                    job.status.state = "completed"
                    job.status.message = "Finalizado"
                job.status.percent = 100.0 if job.status.total_items else 0

            if job.status.state == "stopped":
                job.emit("Execução interrompida.", "AVISO")
            else:
                job.emit("Automação concluída.", "SUCESSO")

        except BaseException as e:
            if isinstance(e, StopRequestedException):
                job.emit("Execução interrompida pelo usuário.", "AVISO")
                with job.lock:
                    job.status.state = "stopped"
                    job.status.message = "Parado pelo usuário"
            else:
                try:
                    job.take_screenshot("erro_critico_execucao")
                except Exception:
                    pass
                job.emit(f"Erro crítico: {e}", "ERRO")
                job.emit(traceback.format_exc(), "DEBUG")
                with job.lock:
                    job.status.state = "error"
                    job.status.message = str(e)
        finally:
            print(f"[JOB {job.status.id}] Iniciando finalização e persistência...")
            video_path = None
            try:
                if job.page and job.page.video:
                    video_path = job.page.video.path()
            except Exception as e:
                print(f"[JOB {job.status.id}] Erro ao obter caminho do vídeo: {e}")

            try:
                self._close_resources(job)
            except Exception as e:
                print(f"[JOB {job.status.id}] Erro ao fechar recursos: {e}")

            try:
                self._collect_video_artifact(job, video_path)
            except Exception as e:
                print(f"[JOB {job.status.id}] Erro ao coletar vídeo: {e}")

            try:
                self._finalize_result_file(job)
            except Exception as e:
                print(f"[JOB {job.status.id}] Erro ao finalizar arquivo de resultado: {e}")

            try:
                self._append_result_history(job)
                print(f"[JOB {job.status.id}] Histórico de resultados persistido.")
            except Exception as e:
                print(f"[JOB {job.status.id}] Erro ao persistir histórico de resultados: {e}")

            try:
                self._append_logs_history(job)
                print(f"[JOB {job.status.id}] Histórico de logs persistido.")
            except Exception as e:
                print(f"[JOB {job.status.id}] Erro ao persistir histórico de logs: {e}")

            with job.condition:
                job.condition.notify_all()
            print(f"[JOB {job.status.id}] Thread de execução finalizada.")


    def _sleep_or_stop(self, job: JobRuntime, seconds: float) -> None:
        end = time.time() + max(seconds, 0)
        while time.time() < end:
            if job.stop_event.is_set():
                return
            while not job.pause_event.is_set():
                if job.stop_event.is_set():
                    return
                time.sleep(0.2)
            time.sleep(0.1)

    def _reset_session(self, job: JobRuntime, nro_cotacao: str, atraso_fases: float) -> None:
        if job.stop_event.is_set():
            return
        job.emit(f"[F0] [Item {nro_cotacao}] Reset de sessão via relogin.", "DEBUG")
        login = job.config.login
        senha = job.config.senha
        job.page = fase2_login.perform_login(job.browser_context, login, senha, URL_DESTINO, job.emit, existing_page=job.page)
        if not job.page:
            raise RuntimeError("Falha no relogin durante reset de sessão")
        self._sleep_or_stop(job, atraso_fases)

    def _is_planilha_tratada(self, filepath: str) -> bool:
        headers_esperados = ["Nro cotação", "Categoria veículo", "Cidade", "UF", "Nome", "Placa", "Data pagamento", "Viagem extra", "Remetente", "Status"]
        path = Path(filepath)
        if path.suffix.lower() == ".xls":
            wb = xlrd.open_workbook(filepath, formatting_info=False)
            sheet = wb.sheet_by_index(0)
            headers_encontrados = sheet.row_values(0) if sheet.nrows > 0 else []
            headers_limpos = [h for h in headers_encontrados if h is not None and str(h).strip() != ""]
            return headers_limpos == headers_esperados

        workbook = openpyxl.load_workbook(filepath, read_only=True)
        try:
            sheet = workbook.active
            headers_encontrados = [cell.value for cell in sheet[1]]
            headers_limpos = [h for h in headers_encontrados if h is not None]
            return headers_limpos == headers_esperados
        finally:
            workbook.close()

    def _ler_dados_planilha(self, job: JobRuntime) -> list[dict]:
        if not job.planilha_processada_path:
            return []
        wb = openpyxl.load_workbook(job.planilha_processada_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        dados = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            row_data = dict(zip(headers, row))
            if row_data.get("Nro cotação"):
                dados.append(row_data)
        wb.close()
        return dados

    def _close_resources(self, job: JobRuntime) -> None:
        """
        Encerra os recursos do navegador Playwright de forma limpa,
        garantindo o salvamento e registro do trace antes de fechar o contexto.
        """
        try:
            if job.browser_context:
                trace_filename = f"{job.log_session_id}_trace.zip"
                trace_filepath = ARTIFACTS_DIR / trace_filename
                try:
                    job.browser_context.tracing.stop(path=str(trace_filepath))
                    job.emit(f"Trace do Playwright salvo: {trace_filename}", "DEBUG")
                    
                    with job.lock:
                        if not any(a.get("name") == trace_filename for a in job.artifacts):
                            job.artifacts.append({
                                "type": "trace",
                                "name": trace_filename,
                                "path": str(trace_filepath.resolve())
                            })
                    self._record_artifact(job.status.id, "trace", trace_filepath)
                except Exception as trace_err:
                    job.emit(f"Falha ao salvar trace do Playwright: {trace_err}", "DEBUG")
        except Exception:
            pass

        try:
            if job.browser_context:
                job.browser_context.close()
        except Exception:
            pass
        try:
            if job.browser:
                job.browser.close()
        except Exception:
            pass

    def _collect_video_artifact(self, job: JobRuntime, video_path: Optional[str] = None) -> None:
        """
        Copia o vídeo gravado pelo Playwright para o diretório de artefatos
        e o registra no banco de dados.
        """
        try:
            source_path = video_path
            if not source_path and job.page and job.page.video:
                try:
                    source_path = job.page.video.path()
                except Exception:
                    pass
            if not source_path:
                return
            source = Path(source_path)
            if not source.exists():
                return
            target = ARTIFACTS_DIR / f"{job.log_session_id}_{source.name}"
            shutil.copyfile(source, target)
            with job.lock:
                # Evita duplicidade se já houver
                if not any(a.get("name") == target.name for a in job.artifacts):
                    job.artifacts.append({"type": "video", "name": target.name, "path": str(target.resolve())})
            self._record_artifact(job.status.id, "video", target)
        except Exception:
            pass
        try:
            if job.playwright_instance:
                job.playwright_instance.stop()
        except Exception:
            pass



manager = AutomationManager()
